import re
import subprocess
import openpyxl
from openpyxl.styles import Font, PatternFill, NamedStyle, Side, Border


def reader(filename):

    users_list = []
    uniq_1 = []
    uniq_2 = []
    uniq_3 = []
    with open(filename, 'r') as f:
        for line in f:
            result = re.split(' ', line, maxsplit=3)
            users_result = [result[0], result[2]]
            users_list.append(users_result)
    for item in users_list:
        count_item = users_list.count(item)
        users_result_count = [item[0], item[1], count_item]
        uniq_1.append(users_result_count)
    for item in uniq_1:
        if item not in uniq_2:
            uniq_2.append(item)
    for item in uniq_2:
        if item[1] == str('Зашел'):
            ad_user_name = item[0]
            cmd = f"Get-ADUser -Identity {ad_user_name} -Properties * | select SamAccountName, Name, Company, Department, Title, employeeType"
            info_from_ad = run(cmd)
            user_and_ad_info = [j for i in [item, info_from_ad] for j in i]
            uniq_3.append(user_and_ad_info)
            print(user_and_ad_info)
    return save_result(uniq_3)


def run(cmd):
    rgx = r"\:\s.*\n"
    subprocess.run(["powershell", "-Command", "chcp 1251"], capture_output=True)
    completed = subprocess.run(["powershell", "-Command", cmd], capture_output=True, text=True)
    subprocess_output = completed.stdout
    fixed_subprocess_output = re.findall(rgx, subprocess_output)
    result_subprocess = []
    for item in fixed_subprocess_output:
        replace_item = item.replace(": ", "").replace('\n', '')
        result_subprocess.append(replace_item)
    return result_subprocess


def rename_column(filename_xlsx):
    book = openpyxl.load_workbook(filename=filename_xlsx)
    sheet : worksheet = book['Users']
    sheet.insert_rows(0)
    sheet['A1'].value = 'Пользователь'
    sheet['B1'].value = 'Статус'
    sheet['C1'].value = 'Количество входов'
    sheet['D1'].value = 'SamAccountName'
    sheet['E1'].value = 'Имя Пользователя'
    sheet['F1'].value = 'Организация'
    sheet['G1'].value = 'Отдел'
    sheet['H1'].value = 'Должность'
    sheet['I1'].value = 'Дирекция'

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 23
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 35
    sheet.column_dimensions['F'].width = 45
    sheet.column_dimensions['G'].width = 60
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 30

    azzcode_style = NamedStyle(name="azzcode_style")
    azzcode_style.font = Font(b=True, size=14, color="DD0000")
    azzcode_style.fill = PatternFill("solid", fgColor="FFFF99")
    side = Side(style='medium', color="00EEDD")
    azzcode_style.border = Border(bottom=side)
    #
    sheet['A1'].style = azzcode_style
    sheet['B1'].style = azzcode_style
    sheet['C1'].style = azzcode_style
    sheet['D1'].style = azzcode_style
    sheet['E1'].style = azzcode_style
    sheet['F1'].style = azzcode_style
    sheet['G1'].style = azzcode_style
    sheet['H1'].style = azzcode_style
    sheet['I1'].style = azzcode_style

    sheet.auto_filter.ref = 'A1:H999'

    book.save(filename_xlsx)


def save_result(usr):

    book = openpyxl.Workbook()
    book.remove(book.active)
    sheet_1 = book.create_sheet('Users')
    for sheet in book.worksheets:
        for row in usr:
            sheet.append(row)

    book.save('./result_output.xlsx')
    return rename_column('./result_output.xlsx')


def main():
    filename = './USERLOG.TXT'
    reader(filename)


if __name__ == "__main__":
    main()
